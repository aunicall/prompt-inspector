"""
Import threat category configs from an Excel file.

Usage:
    python -m scripts.import_category_configs --file <path_to_xlsx>

Excel format:
    Required columns: Domain, Category, Name, Description
    All rows will have severity=critical and enabled=true.
"""

import sys
import asyncio
import argparse
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
from sqlalchemy import delete
from app.database import async_session, init_db
from app.models.category_config import CategoryConfig


async def clear_existing_data():
    async with async_session() as db:
        await db.execute(delete(CategoryConfig))
        await db.commit()
        print("Cleared all existing category_configs data")


async def import_from_excel(excel_path: str):
    if not Path(excel_path).exists():
        print(f"Error: File not found: {excel_path}")
        sys.exit(1)

    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True)
        ws = wb.active
    except Exception as e:
        print(f"Error: Failed to load Excel file: {e}")
        sys.exit(1)

    headers = [cell.value for cell in ws[1]]
    required_columns = ['Domain', 'Category', 'Name', 'Description']
    for col in required_columns:
        if col not in headers:
            print(f"Error: Missing required column: {col}")
            print(f"Found columns: {headers}")
            sys.exit(1)

    col_indices = {col: headers.index(col) for col in required_columns}

    categories = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        try:
            domain = row[col_indices['Domain']]
            category = row[col_indices['Category']]
            name = row[col_indices['Name']]
            description = row[col_indices['Description']]

            if not all([domain, category, name, description]):
                print(f"Warning: Row {row_idx} missing required fields, skipping")
                continue

            categories.append({
                'domain': str(domain).strip(),
                'category': str(category).strip(),
                'name': str(name).strip(),
                'description': str(description).strip(),
                'severity': 'critical',
                'enabled': True,
            })
        except Exception as e:
            print(f"Warning: Row {row_idx} parse error: {e}")
            continue

    wb.close()

    if not categories:
        print("Error: No valid data found in Excel file")
        sys.exit(1)

    print(f"Parsed {len(categories)} categories from Excel")

    await init_db()
    await clear_existing_data()

    async with async_session() as db:
        for cat_data in categories:
            db.add(CategoryConfig(**cat_data))
        await db.commit()
        print(f"Successfully imported {len(categories)} categories")

    print("\nImported categories:")
    for cat in categories:
        print(f"  - [{cat['domain']}] {cat['category']}: {cat['name']}")


def main():
    parser = argparse.ArgumentParser(
        description='Import category configs from Excel file',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Example:
    python -m scripts.import_category_configs --file categories.xlsx
        """
    )
    parser.add_argument('--file', required=True, help='Path to Excel file (.xlsx)')
    args = parser.parse_args()
    asyncio.run(import_from_excel(args.file))


if __name__ == '__main__':
    main()
